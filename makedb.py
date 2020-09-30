import os
import openpyxl
from openpyxl.utils import get_column_letter
import sqlite3
from prettytable import PrettyTable
from _tkinter import create
from openpyxl.utils.cell import _get_column_letter

dbDIR = r"C:/Users/TEMP/Desktop/hour prohectds"

def ptex(ct):
    x = PrettyTable()
    x.field_names = ["City name", "Area", "Population", "Annual Rainfall"]
    x.add_row([ct, 200, 1158259, 600.5])
    x.add_row(["Brisbane", 5905, 1857594, 1146.4])
    x.add_row(["Darwin", 112, 120900, 1714.7])
    x.add_row(["Hobart", 1357, 205556, 619.5])
    x.add_row(["Sydney", 2058, 4336374, 1214.8])
    x.add_row(["Melbourne", 1566, 3806092, 646.9])
    x.add_row([ct, 5386, 1554769, 869.4])
    print(x)
    
fname=os.path.join(dbDIR,"Sales.xlsx")
wb=openpyxl.load_workbook(fname, data_only=True)
shts=wb.sheetnames
sht=wb[shts[0]]
endRow=sht.max_row
endCol=sht.max_column


fname=os.path.join(dbDIR,"sales.db")
conn=sqlite3.connect(fname)
c=conn.cursor()

    
query='DROP TABLE IF EXISTS "'+shts[0]+'"'
print(query)
c.execute(query)
conn.commit()
query='CREATE TABLE IF NOT EXISTS "'+shts[0]+'" (id INTEGER PRIMARY KEY AUTOINCREMENT'
hdrs=[]
for i in range(1,endCol):
    hdrs.append(sht[get_column_letter(i)+"1"].value)
    
for hdr in hdrs:
    query=query+', "'+hdr+'" TEXT'
    
query=query+')'

c.execute(query)
conn.commit()

prehdr=""
for i in range(0,len(hdrs)):
    prehdr=prehdr+'"'+hdrs[i]+'",'
    
prehdr=prehdr[:-1]
for i in range(2,endRow):
    row=[]
    for j in range(1,endCol):
        row.append(sht[get_column_letter(j)+str(i)].value)
    




